from zipfile import BadZipFile

from django.db import connection

from openpyxl import load_workbook

from chartofaccountDIT.models import (
    Analysis1,
    Analysis2,
    NaturalCode,
    ProgrammeCode,
    ProjectCode,
)

from core.import_csv import get_fk, get_fk_from_field, xslx_header_to_dict
from core.models import FinancialYear

from costcentre.models import CostCentre

from forecast.models import (
    Budget,
    FinancialPeriod,
    MonthlyFigure,
    UploadingActuals,
    UploadingBudgets,
)

from upload_file.models import FileUpload
from upload_file.utils import set_file_upload_error

CHART_OF_ACCOUNT_COL = "D"
ACTUAL_FIGURE_COL = "F"
NAC_COL = "A"

TB_FIRST_DATA_ROW = 13

MONTH_CELL = "B2"
TITLE_CELL = "B1"
CORRECT_ACTUAL_TITLE = "Detail Trial Balance"
CORRECT_ACTUAL_WS_NAME = "FNDWRR"

# Sample chart of account entry
# '3000-30000-109189-52191003-310940-00000-00000-0000-0000-0000' # noqa
# The following are the index for the
# chart of account after decoded into a list
CC_INDEX = 2
NAC_INDEX = 3
PROGRAMME_INDEX = 4
ANALYSIS1_INDEX = 5
ANALYSIS2_INDEX = 6
PROJECT_INDEX = 7
CHART_ACCOUNT_SEPARATOR = "-"

VALID_ECONOMIC_CODE_LIST = ['RESOURCE', 'CAPITAL']

# TODO Read the value from the database. It should be
# possible for the business to change it.
GENERIC_PROGRAMME_CODE = 310940


class UploadFileFormatError(Exception):
    pass


class UploadFileDataError(Exception):
    pass


def sql_for_data_copy(data_type):
    if data_type == FileUpload.ACTUALS:
        temp_data_file = 'UploadingActuals'
        target = 'forecast_MonthlyFigure'
    else:
        if data_type == FileUpload.BUDGET:
            temp_data_file = 'UploadingBudgets'
            target = 'Budget'
        else:
            raise UploadFileDataError(
                'Unknown upload type.'
            )

    return 'INSERT INTO {}(' \
           'created, ' \
           'updated, ' \
           'active,  ' \
           'analysis1_code_id, ' \
           'analysis2_code_id, ' \
           'cost_centre_id, ' \
           'financial_period_id, ' \
           'financial_year_id, ' \
           'natural_account_code_id, ' \
           'programme_id, ' \
           'project_code_id, ' \
           'amount, ' \
           'forecast_expenditure_type_id)' \
           ' SELECT  ' \
           'now(), ' \
           'now(), ' \
           'active,  ' \
           'analysis1_code_id, ' \
           'analysis2_code_id, ' \
           'cost_centre_id, ' \
           'financial_period_id, ' \
           'financial_year_id, ' \
           'natural_account_code_id, ' \
           'programme_id, ' \
           'project_code_id, ' \
           'amount, ' \
           'forecast_expenditure_type_id ' \
           ' FROM {};'.format(target, temp_data_file)


def copy_actuals_to_monthly_figure(period_obj, year):
    # Now copy the newly uploaded actuals to the monthly figure table
    MonthlyFigure.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()
    with connection.cursor() as cursor:
        cursor.execute(sql_for_data_copy(FileUpload.ACTUALS))
    UploadingActuals.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()


def get_optional_chart_account_obj(model, chart_of_account_item):
    if int(chart_of_account_item):
        obj, message = get_fk(model, chart_of_account_item)
    else:
        obj = None
        message = ""
    return obj, message


def save_tb_row(chart_of_account, value, period_obj, year_obj):
    """Parse the long strings containing the
    chart of account information. Return errors
    if the elements of the chart of account are missing from database."""
    chart_account_list = chart_of_account.split(CHART_ACCOUNT_SEPARATOR)
    programme_code = chart_account_list[PROGRAMME_INDEX]
    # TODO put GENERIC_PROGRAMME_CODE in database
    # Handle lines without programme code
    if not int(programme_code):
        if value:
            programme_code = GENERIC_PROGRAMME_CODE
        else:
            return True, ""
    cost_centre = chart_account_list[CC_INDEX]
    nac = chart_account_list[NAC_INDEX]
    analisys1 = chart_account_list[ANALYSIS1_INDEX]
    analisys2 = chart_account_list[ANALYSIS2_INDEX]
    project_code = chart_account_list[PROJECT_INDEX]

    error_message = ""
    analysis1_obj = None
    analysis2_obj = None
    project_obj = None
    nac_obj, message = get_fk(NaturalCode, nac)
    error_message += message
    if nac_obj:
        #  Check that the NAC is resource or capital
        if not nac_obj.economic_budget_code or \
                nac_obj.economic_budget_code.upper() not in VALID_ECONOMIC_CODE_LIST:
            return True, ""
    cc_obj, message = get_fk(CostCentre, cost_centre)
    error_message += message
    programme_obj, message = get_fk(ProgrammeCode, programme_code)
    error_message += message
    if analisys1:
        analysis1_obj, message = get_optional_chart_account_obj(
            Analysis1,
            analisys1,
        )
        error_message += message
    if analisys2:
        analysis2_obj, message = get_optional_chart_account_obj(
            Analysis2,
            analisys2,
        )
        error_message += message
    if project_code:
        project_obj, message = get_optional_chart_account_obj(
            ProjectCode,
            project_code,
        )
        error_message += message
    if error_message:
        raise UploadFileDataError(
            error_message
        )

    actuals_obj, created = UploadingActuals.objects.get_or_create(
        financial_year=year_obj,
        programme=programme_obj,
        cost_centre=cc_obj,
        natural_account_code=nac_obj,
        analysis1_code=analysis1_obj,
        analysis2_code=analysis2_obj,
        project_code=project_obj,
        financial_period=period_obj,
    )
    if created:
        # to avoid problems with precision,
        # we store the figures in pence
        actuals_obj.amount = value * 100
    else:
        actuals_obj.amount += value * 100

    actuals_obj.save()
    return True


def check_trial_balance_format(ws, period, year):
    """Check that the file is really the trial
    balance and it is the correct period"""

    if ws[TITLE_CELL].value != CORRECT_ACTUAL_TITLE:
        # wrong file
        raise UploadFileFormatError(
            "This file appears to be corrupt (title is incorrect)"
        )

    report_date = ws[MONTH_CELL].value
    if report_date.year != year:
        # wrong date
        raise UploadFileFormatError(
            "File is for wrong year"
        )

    if report_date.month != period:
        # wrong date
        raise UploadFileFormatError(
            "File is for wrong period"
        )

    return True


def validate_excel_file(file_upload, ws_title):
    try:
        wb = load_workbook(
            file_upload.document_file,
            read_only=True,
        )
    except BadZipFile as ex:
        set_file_upload_error(
            file_upload,
            "The file is not in the correct format (.xlsx)",
            "BadZipFile (user file is not .xlsx)",
        )
        raise ex

    ws = wb.worksheets[0]
    if ws.title != ws_title:
        # wrong file
        raise UploadFileFormatError(
            "File appears to be incorrect: worksheet name is '{}', "
            "expected name is '{}".format(ws.title, ws_title)
        )
    return wb, ws


def upload_trial_balance_report(file_upload, month_number, year):
    try:
        wb, ws = validate_excel_file(file_upload, CORRECT_ACTUAL_WS_NAME)
    except UploadFileFormatError as ex:
        set_file_upload_error(
            file_upload,
            str(ex),
            str(ex),
        )
        raise ex

    try:
        check_trial_balance_format(ws, month_number, year)
    except UploadFileFormatError as ex:
        set_file_upload_error(
            file_upload,
            str(ex),
            str(ex),
        )
        wb.close
        raise ex

    year_obj, msg = get_fk(FinancialYear, year)
    period_obj, msg = get_fk_from_field(
        FinancialPeriod,
        "period_calendar_code",
        month_number)

    # Clear the table used to upload the actuals.
    # The actuals are uploaded to to a temporary storage, and copied
    # to the MonthlyFigure when the upload is completed successfully.
    # This means that we always have a full upload.
    UploadingActuals.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()

    for row in range(TB_FIRST_DATA_ROW, ws.max_row):
        # don't delete this comment: useful for debugging, but it gives a
        # 'too complex error'
        # if not row % 100:
        #     print(row)
        chart_of_account = ws["{}{}".format(CHART_OF_ACCOUNT_COL, row)].value
        if chart_of_account:
            actual = ws["{}{}".format(ACTUAL_FIGURE_COL, row)].value
            # No need to save 0 values, because the data has been cleared
            # before starting the upload
            if actual:
                try:
                    save_tb_row(chart_of_account, actual, period_obj, year_obj)
                except UploadFileDataError as ex:
                    wb.close
                    msg = 'Error at row {}: {}'. \
                        format(row, str(ex))
                    set_file_upload_error(
                        file_upload,
                        msg,
                        msg,
                    )
                    raise ex
        else:
            break

    # Now copy the newly uploaded actuals to the monthly figure table
    copy_actuals_to_monthly_figure(period_obj, year)
    period_obj.actual_loaded = True
    # TODO set all previous month to be Actual Loaded
    period_obj.save()
    wb.close
    return True


EXPECTED_BUDGET_HEADERS = ['cost centre',
                           'natural account',
                           'programme',
                           'analysis',
                           'analysis2',
                           'project',
                           'apr',
                           'may',
                           'jun',
                           'jul',
                           'aug',
                           'sep',
                           'oct',
                           'nov',
                           'dec',
                           'jan',
                           'feb',
                           'mar']


def check_budget_header(header_dict, correct_header):
    error_msg = ''
    correct = True
    for elem in correct_header:
        if elem not in header_dict:
            correct = False
            error_msg += f'Header {elem} not found.'
    if not correct:
        raise UploadFileFormatError(error_msg)


def copy_budget_to_monthly_figure(year, month_dict):
    # Now copy the newly uploaded actuals to the monthly figure table
    for month, period_obj in month_dict.items():
        Budget.objects.filter(
            financial_year=year,
            financial_period=period_obj,
        ).delete()
    with connection.cursor() as cursor:
        cursor.execute(sql_for_data_copy(FileUpload.BUDGET))
    UploadingBudgets.objects.filter(
        financial_year=year,
    ).delete()


def get_id(value, length=0):
    if value:
        if length:
            a = f"{value}"
            return a.zfill(length)
        else:
            return value
    return None


def get_forecast_month_dict():
    """Link the column names in the ADI file with
    the foreign key used in the MonthlyFigure to
    identify the period"""
    q = FinancialPeriod.objects. \
        filter(period_calendar_code__gt=0,
               period_calendar_code__lt=15,
               actual_loaded=False).values("period_short_name")
    period_dict = {}
    for e in q:
        per_obj, msg = get_fk_from_field(
            FinancialPeriod, "period_short_name", e["period_short_name"]
        )
        period_dict[e["period_short_name"].lower()] = per_obj
    return period_dict


def upload_budget(file_upload, year):
    try:
        wb, ws = validate_excel_file(file_upload, "Budgets")
    except UploadFileFormatError as ex:
        set_file_upload_error(
            file_upload,
            str(ex),
            str(ex),
        )
        raise ex
    header_dict = xslx_header_to_dict(ws[1])
    try:
        check_budget_header(header_dict, EXPECTED_BUDGET_HEADERS)
    except UploadFileFormatError as ex:
        set_file_upload_error(
            file_upload,
            str(ex),
            str(ex),
        )
        wb.close
        return False

    year_obj, msg = get_fk(FinancialYear, year)
    month_dict = get_forecast_month_dict()
    # Clear the table used to upload the budgets.
    # The budgets are uploaded to to a temporary storage, and copied
    # when the upload is completed successfully.
    # This means that we always have a full upload.
    UploadingBudgets.objects.filter(
        financial_year=year,
    ).delete()
    for row in range(2, ws.max_row):
        cost_centre = ws["{}{}".format(header_dict["cost centre"], row)].value
        if not cost_centre:
            break
        programme_code = ws["{}{}".format(header_dict["programme"], row)].value
        nac = ws["{}{}".format(header_dict["natural account"], row)].value
        analisys1 = get_id(ws["{}{}".format(header_dict["analysis"], row)].value, 5)
        analisys2 = get_id(ws["{}{}".format(header_dict["analysis2"], row)].value, 5)
        project_code = get_id(ws["{}{}".format(header_dict["project"], row)].value, 4)

        for month, period_obj in month_dict.items():
            period_budget = ws["{}{}".format(header_dict[month.lower()], row)].value
            budget_obj, created = UploadingBudgets.objects.get_or_create(
                financial_year=year_obj,
                programme_id=programme_code,
                cost_centre_id=cost_centre,
                natural_account_code_id=nac,
                analysis1_code_id=analisys1,
                analysis2_code_id=analisys2,
                project_code_id=project_code,
                financial_period=period_obj,
            )
            if created:
                # to avoid problems with precision,
                # we store the figures in pence
                budget_obj.amount = period_budget * 100
            else:
                budget_obj.amount += period_budget * 100
            budget_obj.save()
    copy_budget_to_monthly_figure(year, month_dict)
    return True
