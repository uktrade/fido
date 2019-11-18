from zipfile import BadZipFile

from django.db import connection

from openpyxl import load_workbook

from chartofaccountDIT.models import (
    Analysis1,
    Analysis2,
    NaturalCode,
    ProgrammeCode,
    ProjectCode,
)

from core.import_csv import get_fk, get_fk_from_field
from core.models import FinancialYear

from costcentre.models import CostCentre

from forecast.models import (
    FinancialPeriod,
    MonthlyFigure,
    UploadingActuals,
)

from upload_file.utils import set_file_upload_error

CHART_OF_ACCOUNT_COL = "D"
ACTUAL_FIGURE_COL = "F"
NAC_COL = "A"

FIRST_DATA_ROW = 13

MONTH_CELL = "B2"
TITLE_CELL = "B1"
CORRECT_TITLE = "Detail Trial Balance"
CORRECT_WS_TITLE = "FNDWRR"

# Sample chart of account entry
# '3000-30000-109189-52191003-310940-00000-00000-0000-0000-0000' # noqa
# The following are the index for the
# chart of account after decoded into a list
CC_INDEX = 2
NAC_INDEX = 3
PROGRAMME_INDEX = 4
ANALYSIS1_INDEX = 5
ANALYSIS2_INDEX = 6
PROJECT_INDEX = 7
CHART_ACCOUNT_SEPARATOR = "-"

VALID_ECONOMIC_CODE_LIST = ['RESOURCE', 'CAPITAL']

# TODO Read the value from the database. It should be
# possible for the business to change it.
GENERIC_PROGRAMME_CODE = 310940

COPY_DATA_SQL = 'INSERT INTO forecast_monthlyfigure(' \
                'created, ' \
                'updated, ' \
                'active,  ' \
                'analysis1_code_id, ' \
                'analysis2_code_id, ' \
                'cost_centre_id, ' \
                'financial_period_id, ' \
                'financial_year_id, ' \
                'natural_account_code_id, ' \
                'programme_id, ' \
                'project_code_id, ' \
                'amount, ' \
                'forecast_expenditure_type_id)' \
                ' SELECT  ' \
                'now(), ' \
                'now(), ' \
                'active,  ' \
                'analysis1_code_id, ' \
                'analysis2_code_id, ' \
                'cost_centre_id, ' \
                'financial_period_id, ' \
                'financial_year_id, ' \
                'natural_account_code_id, ' \
                'programme_id, ' \
                'project_code_id, ' \
                'amount, ' \
                'forecast_expenditure_type_id ' \
                ' FROM forecast_uploadingactuals;'


class TrialBalanceError(Exception):
    pass


def copy_actuals_to_monthly_figure(period_obj, year):
    # Now copy the newly uploaded actuals to the monthly figure table
    MonthlyFigure.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()
    with connection.cursor() as cursor:
        cursor.execute(COPY_DATA_SQL)
    UploadingActuals.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()


def get_optional_chart_account_obj(model, chart_of_account_item):
    if int(chart_of_account_item):
        obj, message = get_fk(model, chart_of_account_item)
    else:
        obj = None
        message = ""
    return obj, message


def save_row(chart_of_account, value, period_obj, year_obj):
    """Parse the long strings containing the
    chart of account information. Return errors
    if the elements of the chart of account are missing from database."""
    chart_account_list = chart_of_account.split(CHART_ACCOUNT_SEPARATOR)
    programme_code = chart_account_list[PROGRAMME_INDEX]
    # TODO put GENERIC_PROGRAMME_CODE in database
    # Handle lines without programme code
    if not int(programme_code):
        if value:
            programme_code = GENERIC_PROGRAMME_CODE
        else:
            return True, ""

    error_message = ""
    nac_obj, message = get_fk(NaturalCode, chart_account_list[NAC_INDEX])
    error_message += message
    if nac_obj:
        #  Check that the NAC is resource or capital
        if not nac_obj.economic_budget_code or \
                nac_obj.economic_budget_code.upper() not in VALID_ECONOMIC_CODE_LIST:
            return True, ""
    cc_obj, message = get_fk(CostCentre, chart_account_list[CC_INDEX])
    error_message += message
    programme_obj, message = get_fk(ProgrammeCode, programme_code)
    error_message += message
    analysis1_obj, message = get_optional_chart_account_obj(
        Analysis1,
        chart_account_list[ANALYSIS1_INDEX],
    )
    error_message += message
    analysis2_obj, message = get_optional_chart_account_obj(
        Analysis2,
        chart_account_list[ANALYSIS2_INDEX],
    )
    error_message += message
    project_obj, message = get_optional_chart_account_obj(
        ProjectCode,
        chart_account_list[PROJECT_INDEX],
    )
    error_message += message
    if error_message:
        raise TrialBalanceError(
            error_message
        )

    actuals_obj, created = UploadingActuals.objects.get_or_create(
        financial_year=year_obj,
        programme=programme_obj,
        cost_centre=cc_obj,
        natural_account_code=nac_obj,
        analysis1_code=analysis1_obj,
        analysis2_code=analysis2_obj,
        project_code=project_obj,
        financial_period=period_obj,
    )
    if created:
        # to avoid problems with precision,
        # we store the figures in pence
        actuals_obj.amount = value * 100
    else:
        actuals_obj.amount += value * 100

    actuals_obj.save()

    return True


def check_trial_balance_format(ws, period, year):
    """Check that the file is really the trial
    balance and it is the correct period"""
    if ws.title != CORRECT_WS_TITLE:
        # wrong file
        raise TrialBalanceError(
            "File appears to be incorrect (was it generated by Oracle?)"
        )

    if ws[TITLE_CELL].value != CORRECT_TITLE:
        # wrong file
        raise TrialBalanceError(
            "This file appears to be corrupt (title is incorrect)"
        )

    report_date = ws[MONTH_CELL].value
    if report_date.year != year:
        # wrong date
        raise TrialBalanceError(
            "File is for wrong year"
        )

    if report_date.month != period:
        # wrong date
        raise TrialBalanceError(
            "File is for wrong period"
        )

    return True


def upload_trial_balance_report(file_upload, month_number, year):
    try:
        wb = load_workbook(
            file_upload.document_file,
            read_only=True,
        )
    except BadZipFile as ex:
        set_file_upload_error(
            file_upload,
            "The file is not in the correct format (.xlsx)",
            "BadZipFile (user file is not .xlsx)",
        )
        raise ex

    ws = wb.worksheets[0]

    try:
        check_trial_balance_format(
            ws,
            month_number,
            year,
        )
    except TrialBalanceError as ex:
        wb.close
        set_file_upload_error(
            file_upload,
            str(ex),
            str(ex),
        )
        raise ex

    year_obj, msg = get_fk(FinancialYear, year)
    period_obj, msg = get_fk_from_field(
        FinancialPeriod,
        "period_calendar_code",
        month_number)
    # Clear the table used to upload the actuals.
    # The actuals are uploaded to to a temporary storage, and copied
    # to the MonthlyFigure when the upload is completed successfully.
    # This means that we always have a full upload.

    UploadingActuals.objects.filter(
        financial_year=year,
        financial_period=period_obj,
    ).delete()

    for row in range(FIRST_DATA_ROW, ws.max_row):
        # don't delete this comment: useful for debugging, but it gives a
        # 'too complex error'
        # if not row % 100:
        #     print(row)
        chart_of_account = ws["{}{}".format(CHART_OF_ACCOUNT_COL, row)].value
        if chart_of_account:
            actual = ws["{}{}".format(ACTUAL_FIGURE_COL, row)].value
            # No need to save 0 values, because the data has been cleared
            # before starting the upload
            if actual:
                try:
                    save_row(chart_of_account, actual, period_obj, year_obj)
                except TrialBalanceError as ex:
                    wb.close
                    msg = 'Error at row {}: {}'. \
                        format(row, str(ex))
                    set_file_upload_error(
                        file_upload,
                        msg,
                        msg,
                    )
                    raise ex
        else:
            break

    # Now copy the newly uploaded actuals to the monthly figure table
    copy_actuals_to_monthly_figure(period_obj, year)
    period_obj.actual_loaded = True
    period_obj.save()
    wb.close
    return True
