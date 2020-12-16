import io

from django.contrib.auth.models import Permission
from django.urls import reverse

from openpyxl import load_workbook

from core.test.test_base import BaseTestCase

from end_of_month.test.test_utils import SetFullYearArchive


BUDGET_CELL = "Y2"
GROUP_HEADING_CELL = "A1"
GROUP_CODE_CELL = "B2"
APR_COL = 26
ADJ3_COL = 40


class DownloadForecastHierarchyTest(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)

        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()
        self.archive = SetFullYearArchive()

    def check_workbook(self, content, period):
        file = io.BytesIO(content)
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.active
        year_total = 0
        # Read the forecast/actual values to calculate the year total
        # The total is in the workbook, but reading it returns None
        for col in range(APR_COL, ADJ3_COL + 1):
            year_total += ws.cell(column=col, row=2,).value

        # Check group
        self.assertEqual(ws[GROUP_HEADING_CELL].value, "Group name")
        self.assertEqual(ws[GROUP_CODE_CELL].value, self.archive.group_code)

        self.assertEqual(ws[BUDGET_CELL].value,
                         self.archive.archived_budget[period] / 100)

        self.assertEqual(year_total,
                         self.archive.archived_forecast[period] / 100)

    def dit_download(self, test_period):
        response = self.client.get(
            reverse(
                "export_forecast_data_dit",
                kwargs={"period": test_period}
            )
        )
        self.assertEqual(response.status_code, 200)
        self.check_workbook(response.content, test_period)

    def test_dit_download_apr(self):
        self.dit_download(1)

    def test_dit_download_may(self):
        self.dit_download(2)

    def test_dit_download_jun(self):
        self.dit_download(3)

    def test_dit_download_jul(self):
        self.dit_download(4)

    def test_dit_download_aug(self):
        self.dit_download(5)

    def test_dit_download_sep(self):
        self.dit_download(6)

    def test_dit_download_oct(self):
        self.dit_download(7)

    def test_dit_download_nov(self):
        self.dit_download(8)

    def test_dit_download_dec(self):
        self.dit_download(9)

    def test_dit_download_jan(self):
        self.dit_download(10)

    def test_dit_download_feb(self):
        self.dit_download(11)

    def test_dit_download_mar(self):
        self.dit_download(12)

    def test_dit_download_current(self):
        self.dit_download(0)

    def group_download(self, test_period):
        response = self.client.get(
            reverse(
                "export_forecast_data_group",
                kwargs={
                    'group_code': self.archive.group_code,
                    'period': test_period,
                },
            )
        )
        self.assertEqual(response.status_code, 200)
        self.check_workbook(response.content, test_period)

    def test_group_download_apr(self):
        self.group_download(1)

    def test_group_download_may(self):
        self.group_download(2)

    def test_group_download_jun(self):
        self.group_download(3)

    def test_group_download_jul(self):
        self.group_download(4)

    def test_group_download_aug(self):
        self.group_download(5)

    def test_group_download_sep(self):
        self.group_download(6)

    def test_group_download_oct(self):
        self.group_download(7)

    def test_group_download_nov(self):
        self.group_download(8)

    def test_group_download_dec(self):
        self.group_download(9)

    def test_group_download_jan(self):
        self.group_download(10)

    def test_group_download_feb(self):
        self.group_download(11)

    def test_group_download_mar(self):
        self.group_download(12)

    def test_group_download_current(self):
        self.group_download(0)

    def directorate_download(self, test_period):
        response = self.client.get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    'directorate_code': self.archive.directorate_code,
                    'period': test_period,
                },
            )
        )
        self.assertEqual(response.status_code, 200)
        self.check_workbook(response.content, test_period)

    def test_directorate_download_apr(self):
        self.directorate_download(1)

    def test_directorate_download_may(self):
        self.directorate_download(2)

    def test_directorate_download_jun(self):
        self.directorate_download(3)

    def test_directorate_download_jul(self):
        self.directorate_download(4)

    def test_directorate_download_aug(self):
        self.directorate_download(5)

    def test_directorate_download_sep(self):
        self.directorate_download(6)

    def test_directorate_download_oct(self):
        self.directorate_download(7)

    def test_directorate_download_nov(self):
        self.directorate_download(8)

    def test_directorate_download_dec(self):
        self.directorate_download(9)

    def test_directorate_download_jan(self):
        self.directorate_download(10)

    def test_directorate_download_feb(self):
        self.directorate_download(11)

    def test_directorate_download_mar(self):
        self.directorate_download(12)

    def test_directorate_download_current(self):
        self.directorate_download(0)

    def cost_centre_download(self, test_period):
        response = self.client.get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.archive.cost_centre_code,
                    'period': test_period,
                },
            )
        )
        self.assertEqual(response.status_code, 200)
        self.check_workbook(response.content, test_period)

    def test_cost_centre_download_apr(self):
        self.cost_centre_download(1)

    def test_cost_centre_download_may(self):
        self.cost_centre_download(2)

    def test_cost_centre_download_jun(self):
        self.cost_centre_download(3)

    def test_cost_centre_download_jul(self):
        self.cost_centre_download(4)

    def test_cost_centre_download_aug(self):
        self.cost_centre_download(5)

    def test_cost_centre_download_sep(self):
        self.cost_centre_download(6)

    def test_cost_centre_download_oct(self):
        self.cost_centre_download(7)

    def test_cost_centre_download_nov(self):
        self.cost_centre_download(8)

    def test_cost_centre_download_dec(self):
        self.cost_centre_download(9)

    def test_cost_centre_download_jan(self):
        self.cost_centre_download(10)

    def test_cost_centre_download_feb(self):
        self.cost_centre_download(11)

    def test_cost_centre_download_mar(self):
        self.cost_centre_download(12)

    def test_cost_centre_download_current(self):
        self.cost_centre_download(0)
