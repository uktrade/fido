import io

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.myutils import get_current_financial_year
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.test.test_utils import create_budget
from forecast.views.view_forecast.export_forecast_data import (
    export_forecast_data_project_detail_cost_centre,
    export_forecast_data_project_detail_directorate,
    export_forecast_data_project_detail_dit,
    export_forecast_data_project_detail_group,
)


class DownloadProjectDetailyTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)

        self.group_name = "Test Group"
        self.group_code = "TestGG"
        self.directorate_name = "Test Directorate"
        self.directorate_code = "TestDD"
        self.cost_centre_code = 109076

        group_obj = DepartmentalGroupFactory(
            group_code=self.group_code, group_name=self.group_name,
        )
        self.directorate = DirectorateFactory(
            directorate_code=self.directorate_code,
            directorate_name=self.directorate_name,
            group=group_obj,
        )
        cost_centre = CostCentreFactory(
            directorate=self.directorate, cost_centre_code=self.cost_centre_code,
        )
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.project_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        financial_code_obj = FinancialCode.objects.create(
            programme=programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(codename="can_view_forecasts")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.budget = create_budget(financial_code_obj, year_obj)
        self.year_total = self.amount_apr + self.amount_may
        self.underspend_total = self.budget - self.amount_apr - self.amount_may
        self.spend_to_date_total = self.amount_apr

    def check_response_content(self, content):
        file = io.BytesIO(content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code
        assert ws["W1"].value == "Project code"
        assert ws["W2"].value == self.project_code

    def test_dit_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_dit",
                kwargs={"project_code_id": self.project_code, "period": 0},
            ),
            export_forecast_data_project_detail_dit,
            project_code_id=self.project_code,
            period=0,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_group_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_group",
                kwargs={
                    "group_code": self.group_code,
                    "project_code_id": self.project_code,
                    "period": 0,
                },
            ),
            export_forecast_data_project_detail_group,
            group_code=self.group_code,
            project_code_id=self.project_code,
            period=0,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_directorate_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_directorate",
                kwargs={
                    "directorate_code": self.directorate.directorate_code,
                    "project_code_id": self.project_code,
                    "period": 0,
                },
            ),
            export_forecast_data_project_detail_directorate,
            directorate_code=self.directorate.directorate_code,
            project_code_id=self.project_code,
            period=0,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "project_code_id": self.project_code,
                    "period": 0,
                },
            ),
            export_forecast_data_project_detail_cost_centre,
            cost_centre=self.cost_centre_code,
            project_code_id=self.project_code,
            period=0,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)
