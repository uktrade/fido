import io

from django.contrib.auth.models import Permission
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.test.test_base import BaseTestCase
from core.utils.generic_helpers import get_current_financial_year

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.permission_shortcuts import assign_perm
from forecast.test.test_utils import create_budget


class DownloadForecastHierarchyTest(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.group_name = "Test Group"
        self.group_code = "TestGG"
        self.directorate_name = "Test Directorate"
        self.directorate_code = "TestDD"
        self.cost_centre_code = 109076

        self.group = DepartmentalGroupFactory(
            group_code=self.group_code,
            group_name=self.group_name,
        )
        self.directorate = DirectorateFactory(
            directorate_code=self.directorate_code,
            directorate_name=self.directorate_name,
            group=self.group,
        )
        self.cost_centre = CostCentreFactory(
            directorate=self.directorate,
            cost_centre_code=self.cost_centre_code,
        )
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        self.project_obj = ProjectCodeFactory()
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=nac_obj,
            project_code=self.project_obj
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=1
            ),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=2,
            ),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj
        )
        may_figure.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.budget = create_budget(financial_code_obj, year_obj)
        self.year_total = self.amount_apr + self.amount_may
        self.underspend_total = self.budget - self.amount_apr - self.amount_may
        self.spend_to_date_total = self.amount_apr

    def test_dit_download(self):
        dit_url = self.client.get(
            reverse(
                "export_forecast_data_dit",
                kwargs={"period": 0}
            )
        )

        self.assertEqual(dit_url.status_code, 200)

        file = io.BytesIO(dit_url.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code

    def test_dit_cannot_download(self):
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.remove(can_view_forecasts)

        dit_url = self.client.get(
            reverse(
                "export_forecast_data_dit",
                kwargs={'period': 0}
            )
        )

        self.assertEqual(dit_url.status_code, 302)

    def test_group_download(self):
        response = self.client.get(
            reverse(
                "export_forecast_data_group",
                kwargs={
                    'group_code': self.group.group_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code

    def test_group_cannot_download(self):
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.remove(can_view_forecasts)

        response = self.client.get(
            reverse(
                "export_forecast_data_group",
                kwargs={
                    'group_code': self.group.group_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 302)

    def test_directorate_download(self):
        response = self.client.get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    'directorate_code': self.directorate.directorate_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code

    def test_directorate_cannot_download(self):
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.remove(can_view_forecasts)

        response = self.client.get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    'directorate_code': self.directorate.directorate_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 302)

    def test_cost_centre_download(self):
        assign_perm("change_costcentre", self.test_user, self.cost_centre)
        self.cost_centre_code = self.cost_centre

        response = self.client.get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre.cost_centre_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code

    def test_cost_centre_cannot_download(self):
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.remove(can_view_forecasts)

        response = self.client.get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre.cost_centre_code,
                    'period': 0,
                },
            )
        )

        self.assertEqual(response.status_code, 302)
