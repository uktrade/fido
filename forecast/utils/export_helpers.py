from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
)

from core.exportutils import (
    EXCEL_TYPE,
    EXC_TAB_NAME_LEN,
)
from core.utils import today_string

from forecast.models import FinancialPeriod

BUDGET_HEADER = 'Budget'


def format_numbers(ws, row, start):
    for c in range(start, start + 16):
        ws[f'{get_column_letter(c)}{row}'].number_format = '#,##0.00'


def unlock_forecast_cells(ws, row, start, end):
    for c in range(start, end):
        ws[f'{get_column_letter(c)}{row}'].protection = Protection(locked=False)


def forecast_query_iterator(queryset, keys_dict, columns_dict, period_list):
    for obj in queryset:
        row = []
        for field in keys_dict.keys():
            val = obj[field]
            if val is None:
                val = ""
            row.append(val)
        row.append(obj['Budget'] / 100)
        for period in period_list:
            row.append(obj[period] / 100)
        row.append('')
        row.append('')
        row.append('')
        for field in columns_dict.keys():
            val = obj[field]
            if val is None:
                val = ""
            row.append(val)
        yield row


def create_headers(keys_dict, columns_dict, period_list):
    k = list(keys_dict.values())
    k.append(BUDGET_HEADER)
    k.extend(period_list)
    k.append('Year to Date')
    k.append('Year Total')
    k.append('Underspend/Overspend')
    k.extend(list(columns_dict.values()))
    return k


def last_actual_cell(col):
    actual_month = FinancialPeriod.financial_period_info.actual_month() - 1
    if actual_month:
        return (get_column_letter(column_index_from_string(col) + actual_month))
    return 0


def export_to_excel(queryset,
                    columns_dict,
                    extra_columns_dict,
                    protect,
                    title):
    resp = HttpResponse(content_type=EXCEL_TYPE)
    filename = f'{title}  {today_string()} .xlsx'
    resp["Content-Disposition"] = "attachment; filename=" + filename
    wb = Workbook()
    ws = wb.get_active_sheet()
    # Truncate the tab name to the maximum lenght permitted by Excel
    ws.title = f'{title} {today_string()}'[:EXC_TAB_NAME_LEN]
    if protect:
        #  Set the required level of protection
        ws.protection.sheet = True
        ws.protection.autoFilter = False
        ws.protection.sort = False
        ws.protection.pivotTables = False
        ws.protection.formatCells = False
        ws.protection.formatRows = False
        ws.protection.formatColumns = False
    row_count = 1
    period_list = FinancialPeriod.financial_period_info.period_display_list()
    howmany_periods = len(period_list)
    header = create_headers(columns_dict, extra_columns_dict, period_list)
    budget_index = header.index(BUDGET_HEADER) + 1
    budget_col = get_column_letter(budget_index)
    first_actual_col = get_column_letter(budget_index + 1)
    last_actual_col = last_actual_cell(first_actual_col)
    if last_actual_col:
        first_forecast_index = column_index_from_string(last_actual_col) + 1
    else:
        first_forecast_index = budget_index + 1
    last_month_index = budget_index + howmany_periods
    last_month_col = get_column_letter(last_month_index)
    year_to_date_col = get_column_letter(last_month_index + 1)
    year_total_col = get_column_letter(last_month_index + 2)
    over_under_spend_col = get_column_letter(last_month_index + 3)
    ws.append(header)

    for data_row in forecast_query_iterator(
            queryset,
            columns_dict,
            extra_columns_dict,
            period_list
    ):
        ws.append(data_row)
        row_count += 1
        # Formula for Year To Date. Don't use it if there are no actuals
        if last_actual_col:
            ws[f'{year_to_date_col}{row_count}'].value = \
                f'=SUM({first_actual_col}{row_count}:{last_actual_col}{row_count})'
        else:
            ws[f'{year_to_date_col}{row_count}'].value = 0

        # Formula for calculating the full year
        ws[f'{year_total_col}{row_count}'].value = \
            f'=SUM({first_actual_col}{row_count}:{last_month_col}{row_count})'
        ws[f'{over_under_spend_col}{row_count}'].value = \
            f'=({budget_col}{row_count}-{year_to_date_col}{row_count})'
        format_numbers(ws, row_count, budget_index)
        unlock_forecast_cells(ws, row_count, first_forecast_index, last_month_index)

    wb.save(resp)
    return resp


def export_query_to_excel(queryset, columns_dict, title):
    return export_to_excel(queryset,
                           columns_dict,
                           {},
                           False,
                           title
                           )


def export_edit_to_excel(queryset, key_dict, columns_dict, title):
    return export_to_excel(queryset,
                           key_dict,
                           columns_dict,
                           True,
                           title
                           )
