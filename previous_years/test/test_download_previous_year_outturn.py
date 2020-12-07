import io

from django.urls import reverse

from openpyxl import load_workbook

from forecast.views.view_forecast.export_forecast_data import (
    export_forecast_data_cost_centre,
    export_forecast_data_directorate,
    export_forecast_data_dit,
    export_forecast_data_expenditure_detail_cost_centre,
    export_forecast_data_expenditure_detail_directorate,
    export_forecast_data_expenditure_detail_group,
    export_forecast_data_expenditure_dit,
    export_forecast_data_group,
    export_forecast_data_programme_detail_directorate,
    export_forecast_data_programme_detail_dit,
    export_forecast_data_programme_detail_group,
    export_forecast_data_project_detail_cost_centre,
    export_forecast_data_project_detail_directorate,
    export_forecast_data_project_detail_dit,
    export_forecast_data_project_detail_group,
)

from previous_years.test.test_utils import (
    PastYearForecastSetup,
    hide_adjustment_columns,
)


class DownloadPastYearForecastTest(PastYearForecastSetup):
    def check_response_content(self, content):
        file = io.BytesIO(content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["B1"].value == "Group code"
        assert ws["B2"].value == self.group_code
        assert ws["D1"].value == "Directorate code"
        assert ws["D2"].value == self.directorate_code
        assert ws["F1"].value == "Cost Centre code"
        assert ws["F2"].value == self.cost_centre_code

        assert ws["N1"].value == "PO/Actual NAC"
        assert ws["N2"].value == self.natural_account_code
        assert ws["Q1"].value == "Programme code"
        assert ws["Q2"].value == self.programme_code
        assert ws["W1"].value == "Project code"
        assert ws["W2"].value == self.project_code
        assert ws["S1"].value == "Contract code"
        assert ws["S2"].value == self.analisys1
        assert ws["U1"].value == "Market code"
        assert ws["U2"].value == self.analisys2
        assert ws["J1"].value == "Budget Type"
        assert ws["J2"].value == self.budget_type_id

        # print(f'{ws["G2"].value} {ws["H2"].value} {ws["J2"].value}')
        # check the figures
        assert ws["Y2"].value == self.outturn["budget"]
        assert ws["Z2"].value == self.outturn["apr"]
        assert ws["AA2"].value == self.outturn["may"]
        assert ws["AB2"].value == self.outturn["jun"]
        assert ws["AC2"].value == self.outturn["jul"]
        assert ws["AD2"].value == self.outturn["aug"]
        assert ws["AE2"].value == self.outturn["sep"]
        assert ws["AF2"].value == self.outturn["oct"]
        assert ws["AG2"].value == self.outturn["nov"]
        assert ws["AH2"].value == self.outturn["dec"]
        assert ws["AI2"].value == self.outturn["jan"]
        assert ws["AJ2"].value == self.outturn["feb"]
        assert ws["AK2"].value == self.outturn["mar"]
        assert ws["AL2"].value == self.outturn["adj01"]
        assert ws["AM2"].value == self.outturn["adj02"]
        assert ws["AN2"].value == self.outturn["adj03"]

    def test_dit_download(self):
        response = self.factory_get(
            reverse("export_forecast_data_dit", kwargs={"period": self.archived_year}),
            export_forecast_data_dit,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_group_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_group",
                kwargs={"group_code": self.group_code, "period": self.archived_year, },
            ),
            export_forecast_data_group,
            group_code=self.group_code,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_directorate_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_directorate,
            directorate_code=self.directorate_code,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_cost_centre,
            cost_centre=self.cost_centre_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_dit_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_dit",
                kwargs={
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_dit,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_group_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_group",
                kwargs={
                    "group_code": self.group_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_group,
            group_code=self.group_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_directorate_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_directorate,
            directorate_code=self.directorate_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_cost_centre,
            cost_centre=self.cost_centre_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_directorate_programme_download(self):
        response = self.factory_get(
            reverse(
                "export_programme_details_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "programme_code_id": self.project_code,
                    "forecast_expenditure_type_name": self.expenditure_type_name,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_programme_detail_directorate,
            directorate_code=self.directorate_code,
            programme_code_id=self.programme_code,
            forecast_expenditure_type_name=self.expenditure_type_name,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_group_programme_download(self):
        response = self.factory_get(
            reverse(
                "export_programme_details_group",
                kwargs={
                    "group_code": self.group_code,
                    "programme_code_id": self.project_code,
                    "forecast_expenditure_type_name": self.expenditure_type_name,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_programme_detail_group,
            group_code=self.group_code,
            programme_code_id=self.programme_code,
            forecast_expenditure_type_name=self.expenditure_type_name,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_dit_programme_download(self):
        response = self.factory_get(
            reverse(
                "export_programme_details_dit",
                kwargs={
                    "programme_code_id": self.project_code,
                    "forecast_expenditure_type_name": self.expenditure_type_name,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_programme_detail_dit,
            programme_code_id=self.programme_code,
            forecast_expenditure_type_name=self.expenditure_type_name,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_expenditure_download(self):
        response = self.factory_get(
            reverse(
                "export_expenditure_details_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "expenditure_category_id": self.expenditure_category_id,
                    "budget_type_id": self.budget_type_id,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_expenditure_detail_cost_centre,
            cost_centre=self.cost_centre_code,
            expenditure_category_id=self.expenditure_category_id,
            budget_type_id=self.budget_type_id,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_directorate_expenditure_download(self):
        response = self.factory_get(
            reverse(
                "export_expenditure_details_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "expenditure_category_id": self.expenditure_category_id,
                    "budget_type_id": self.budget_type_id,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_expenditure_detail_directorate,
            directorate_code=self.directorate_code,
            expenditure_category_id=self.expenditure_category_id,
            budget_type_id=self.budget_type_id,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_group_expenditure_download(self):
        response = self.factory_get(
            reverse(
                "export_expenditure_details_group",
                kwargs={
                    "group_code": self.group_code,
                    "expenditure_category_id": self.expenditure_category_id,
                    "budget_type_id": self.budget_type_id,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_expenditure_detail_group,
            group_code=self.group_code,
            expenditure_category_id=self.expenditure_category_id,
            budget_type_id=self.budget_type_id,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_dit_expenditure_download(self):
        response = self.factory_get(
            reverse(
                "export_expenditure_details_dit",
                kwargs={
                    "expenditure_category_id": self.expenditure_category_id,
                    "budget_type_id": self.budget_type_id,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_expenditure_dit,
            expenditure_category_id=self.expenditure_category_id,
            budget_type_id=self.budget_type_id,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)


class DownloadPastYearForecastAdjustmentColumnsTest(DownloadPastYearForecastTest):
    def setUp(self):
        super().setUp()
        hide_adjustment_columns()
