import os

from django.test import (
    TestCase,
)

from openpyxl import Workbook

from chartofaccountDIT.test.factories import (
    HistoricalAnalysis1Factory,
    HistoricalAnalysis2Factory,
    HistoricalNaturalCodeFactory,
    HistoricalProgrammeCodeFactory,
    HistoricalProjectCodeFactory,
)

from core.models import FinancialYear

from costcentre.test.factories import ArchivedCostCentreFactory

from previous_years.import_previous_year import (
    ANALYSIS2_HEADER,
    ANALYSIS_HEADER,
    ArchiveYearError,
    COST_CENTRE_HEADER,
    DATA_HEADERS,
    NAC_HEADER,
    PROGRAMME_HEADER,
    PROJECT_HEADER,
    VALID_WS_NAME,
    upload_previous_year,
)
from previous_years.models import (
    ArchivedFinancialCode,
    ArchivedForecastData,
)

from upload_file.models import FileUpload


class ImportPreviousYearForecastTest(TestCase):
    def setUp(self):
        # 2019 is created when the database is created, so it exists
        self.archived_year = 2019
        self.archived_year_obj = FinancialYear.objects.get(pk=self.archived_year)
        self.cost_centre_code = "109189"
        self.natural_account_code = 52191003
        self.programme_code = "310940"
        self.project_code = "0123"
        self.analisys1 = "00798"
        self.analisys2 = "00321"
        ArchivedCostCentreFactory.create(
            cost_centre_code=self.cost_centre_code,
            financial_year=self.archived_year_obj,
        )
        HistoricalProjectCodeFactory.create(
            project_code=self.project_code, financial_year=self.archived_year_obj
        )
        HistoricalProgrammeCodeFactory.create(
            programme_code=self.programme_code, financial_year=self.archived_year_obj
        )
        HistoricalNaturalCodeFactory.create(
            natural_account_code=self.natural_account_code,
            economic_budget_code="CAPITAL",
            financial_year=self.archived_year_obj,
        )
        HistoricalAnalysis2Factory.create(
            analysis2_code=self.analisys2, financial_year=self.archived_year_obj
        )
        HistoricalAnalysis1Factory.create(
            analysis1_code=self.analisys1, financial_year=self.archived_year_obj
        )
        self.create_workbook()

    def tearDown(self):
        if os.path.exists(self.excel_file_name):
            os.remove(self.excel_file_name)

    def create_workbook(self):
        wb = Workbook()
        self.data_worksheet = wb.active
        self.data_worksheet.title = VALID_WS_NAME
        col_index = 1
        self.data_worksheet.cell(column=col_index, row=1, value=COST_CENTRE_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.cost_centre_code)
        self.data_worksheet.cell(column=col_index, row=3, value=self.cost_centre_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=NAC_HEADER)
        self.data_worksheet.cell(
            column=col_index, row=2, value=self.natural_account_code
        )
        self.data_worksheet.cell(
            column=col_index, row=3, value=self.natural_account_code
        )
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=PROGRAMME_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.programme_code)
        self.data_worksheet.cell(column=col_index, row=3, value=self.programme_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=PROJECT_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.project_code)
        self.data_worksheet.cell(column=col_index, row=3, value=self.project_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=ANALYSIS_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.analisys1)
        self.data_worksheet.cell(column=col_index, row=3, value=self.analisys1)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=ANALYSIS2_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.analisys2)
        self.data_worksheet.cell(column=col_index, row=3, value=self.analisys2)
        self.results = []
        for month in DATA_HEADERS:
            col_index += 1
            self.data_worksheet.cell(column=col_index, row=1, value=month)
            value1 = col_index * 5
            value2 = col_index * 7
            self.data_worksheet.cell(column=col_index, row=2, value=value1)
            self.data_worksheet.cell(column=col_index, row=3, value=value2)
            self.results.append((value1 + value2) * 100)
        self.excel_file_name = os.path.join(os.path.dirname(__file__), 'dummy.xlsx', )
        wb.save(filename=self.excel_file_name)

    def test_upload_wrong(self):
        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PREVIOUSYEAR,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        with self.assertRaises(ArchiveYearError):
            upload_previous_year(
                self.data_worksheet, self.archived_year + 1, file_upload_obj,
            )

    def test_upload(self):
        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 0)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 0)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PREVIOUSYEAR,
            file_location=FileUpload.LOCALFILE,
        )
        financial_year_obj = FinancialYear.objects.get(pk=self.archived_year)
        self.assertEqual(financial_year_obj.archived, False)
        file_upload_obj.save()
        upload_previous_year(self.data_worksheet, self.archived_year, file_upload_obj,)

        financial_year_obj = FinancialYear.objects.get(pk=self.archived_year)
        self.assertEqual(financial_year_obj.archived, True)

        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 1)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 1)
        result_obj = ArchivedForecastData.objects.all().first()
        self.assertEqual(self.results[0], result_obj.budget)
        self.assertEqual(self.results[1], result_obj.apr)
        self.assertEqual(self.results[2], result_obj.may)
        self.assertEqual(self.results[3], result_obj.jun)
        self.assertEqual(self.results[4], result_obj.jul)
        self.assertEqual(self.results[5], result_obj.aug)
        self.assertEqual(self.results[6], result_obj.sep)
        self.assertEqual(self.results[7], result_obj.oct)
        self.assertEqual(self.results[8], result_obj.nov)
        self.assertEqual(self.results[9], result_obj.dec)
        self.assertEqual(self.results[10], result_obj.jan)
        self.assertEqual(self.results[11], result_obj.feb)
        self.assertEqual(self.results[12], result_obj.mar)
        self.assertEqual(self.results[13], result_obj.adj1)
        self.assertEqual(self.results[14], result_obj.adj2)
        self.assertEqual(self.results[15], result_obj.adj3)
