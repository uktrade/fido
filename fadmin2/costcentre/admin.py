from django.contrib import admin

# Register your models here.
from django.db import models
from django.http import HttpResponse
import csv
from django.http import StreamingHttpResponse
#from core.utils import SmartExport
from .models import DepartmentalGroup, Directorate, CostCentre, Programme

import openpyxl

from openpyxl.utils import get_column_letter



# http://blog.aeguana.com/2015/12/12/csv-export-data-for-django-model/


#  https://djangotricks.blogspot.co.uk/2013/12/how-to-export-data-as-excel.html
# def export_cc_csv(modeladmin, request, queryset):
#
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename=costCentre.csv'
#     writer = csv.writer(response, csv.excel)
#     response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
#     writer.writerow([
#         smart_str(u'Cost Centre'),
#         smart_str(u'Cost Centre Description'),
#         smart_str(u'Directorate Code'),
#         smart_str(u'Directorate Name'),
#         smart_str(u'Departmental Group Code'),
#         smart_str(u'Departmental Group Name')
#     ])
#     for obj in queryset:
#         writer.writerow([
#             smart_str(obj.CCCode),
#             smart_str(obj.CCName),
#             smart_str(obj.Directorate.DirectorateCode),
#             smart_str(obj.Directorate.DirectorateName),
#             smart_str(obj.Directorate.GroupCode.GroupCode),
#             smart_str(obj.Directorate.GroupCode.GroupName)
#         ])
#     return response


# def export_cc1_csv(modeladmin, request, queryset):
#
#     from django.utils.encoding import smart_str
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename=costCentre.csv'
#     writer = csv.writer(response, csv.excel)
#     response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
#
#     scv = SmartExport(queryset)
#
#     for row in scv.stream():
#         writer.writerow(row)
#     return response
#
#
# export_cc1_csv.short_description = u"ShortExport CSV"

# writer = csv.writer(pseudo_buffer)
# response = StreamingHttpResponse(
#     (writer.writerow(row) for row in stream(headers, mydata_qs)),
#     content_type="text/csv")
# response['Content-Disposition'] = 'attachment; filename="all_kittens.csv"'
# return response


# EXPORT_ITERATOR_HEADERS=['Cost Centre','Cost Centre2']
# def _export_iterator(queryset):
#     for obj in queryset:
#         yield(            obj.CCCode,
#             obj.CCName,
#             obj.Directorate.DirectorateCode,
#             obj.Directorate.DirectorateName,
#             obj.Directorate.GroupCode.GroupCode,
#             obj.Directorate.GroupCode.GroupName)


# def export_cc_xlsx(modeladmin, request, queryset):
#     import openpyxl
#     from openpyxl.utils import get_column_letter
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
#     wb = openpyxl.Workbook()
#     ws = wb.get_active_sheet()
#     ws.title = "MyModel"
#
#     row_num = 0
#
#     columns = [
#          (u'Cost Centre', 10),
#          (u'Cost Centre Description', 50),
#          (u'Directorate Code', 10),
#          (u'Directorate Name', 50),
#          (u'Departmental Group Code', 10),
#          (u'Departmental Group Name', 50)
#     ]
#
#     for col_num in range(len(columns)):
#         c = ws.cell(row=row_num + 1, column=col_num + 1)
#         c.value = columns[col_num][0]
#         # c.style.font.bold = True
#         # set column width
#         ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
#
#     for obj in queryset:
#         row_num += 1
#         row = [
#             obj.CCCode,
#             obj.CCName,
#             obj.Directorate.DirectorateCode,
#             obj.Directorate.DirectorateName,
#             obj.Directorate.GroupCode.GroupCode,
#             obj.Directorate.GroupCode.GroupName
#         ]
#         for col_num in range(len(row)):
#             c = ws.cell(row=row_num + 1, column=col_num + 1)
#             c.value = row[col_num]
#
#     wb.save(response)
#     return response

# def export_cc_xlsx(modeladmin, request, queryset):
#
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
#     wb = openpyxl.Workbook()
#     ws = wb.get_active_sheet()
#     ws.title = "MyModel"
#
#     row_num = 0
#     scv = SmartExport(queryset)
#
#     for row in scv.stream():
#         row_num += 1
#         for col_num in range(len(row)):
#             c = ws.cell(row=row_num + 1, column=col_num + 1)
#             c.value = row[col_num]
#
#     wb.save(response)
#     return response
#
#
# export_cc_xlsx.short_description = u"Export XLSX"


# Displays extra fields in the list of cost centres
class CostCentreAdmin(admin.ModelAdmin):
    list_display = ('cost_centre_code', 'cost_centre_name', 'directorate_name', 'dg_name')

    def directorate_name(self, instance): # required to display the filed from a foreign key
        return instance.directorate.directorate_name

    def dg_name(self, instance):
        return instance.directorate.group_code.group_name

    directorate_name.admin_order_field = 'directorate__directorate_name'  # use __ to define a table field relationship
    dg_name.admin_order_field = 'directorate__group_code__group_name'  # use __ to define a table field relationship

    search_fields = ['cost_centre_code']
    list_filter = ['directorate__directorate_name','directorate__group_code__group_name']
    readonly_fields = ['cost_centre_code', 'created', 'updated'] # don't allow to edit the code
    fields = ['cost_centre_code', 'cost_centre_name', 'directorate', 'created', 'updated']  # required to display the read only field at the top
  #  actions = [export_cc1_csv, export_cc_xlsx] # new action to export to csv and xlsx


class DirectorateAdmin(admin.ModelAdmin):
    list_display = ('directorate_name','dgroup_name')

    def dgroup_name(self, instance):
        return instance.group_code.group_name

    dgroup_name.admin_order_field = 'group_code__group_name' # use __ to define a table field relationship


admin.site.register(CostCentre, CostCentreAdmin)
admin.site.register(DepartmentalGroup)
admin.site.register(Directorate, DirectorateAdmin)

admin.site.register(Programme)