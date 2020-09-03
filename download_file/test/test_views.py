import io

from django.contrib.auth.models import Permission
from django.core.exceptions import PermissionDenied
from django.test import TestCase
from django.urls import reverse

from download_file.views.mi_report_download import DownloadMIReportView
from download_file.views.oscar_return import DownloadOscarReturnView

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.test.test_base import RequestFactoryBase
from core.utils.generic_helpers import get_current_financial_year


from costcentre.test.factories import CostCentreFactory

from forecast.models import (
    BudgetMonthlyFigure,
    FinancialCode,
    FinancialPeriod,
)
from forecast.views.export.mi_report_source import export_mi_budget_report


class DownloadViewTests(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)
        self.cost_centre_code = 109076
        self.cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -234567
        self.amount_may = 345216
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.programme_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save
        may_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save
        self.year_total = self.amount_apr + self.amount_may

    def test_download_mi_view(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")

        downloaded_files_url = reverse("download_mi_report",)

        # Should have been redirected (no permission)
        with self.assertRaises(PermissionDenied):
            self.factory_get(
                downloaded_files_url, DownloadMIReportView,
            )

        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        resp = self.factory_get(downloaded_files_url, DownloadMIReportView,)
        # Should have been permission now
        self.assertEqual(resp.status_code, 200)

    def test_download_mi_budget(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")
        downloaded_files_url = reverse("download_mi_budget",)

        response = self.factory_get(downloaded_files_url, export_mi_budget_report,)
        self.assertEqual(response.status_code, 302)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()
        response = self.factory_get(downloaded_files_url, export_mi_budget_report,)
        self.assertEqual(response.status_code, 200)
        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True,)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Entity")
        self.assertEqual(ws["A2"].value, "3000")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["H2"].value, self.amount_apr / 100)
        self.assertEqual(ws["I2"].value, self.amount_may / 100)
        self.assertEqual(ws["W2"].value, self.year_total / 100)
        wb.close()

    def test_download_oscar_view(self):
        assert not self.test_user.has_perm("forecast.can_download_oscar")

        downloaded_files_url = reverse("download_oscar_report",)

        # Should have been redirected (no permission)
        with self.assertRaises(PermissionDenied):
            self.factory_get(
                downloaded_files_url, DownloadOscarReturnView,
            )

        can_download_files = Permission.objects.get(codename="can_download_oscar",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        resp = self.factory_get(downloaded_files_url, DownloadOscarReturnView,)

        # Should have been permission now
        self.assertEqual(resp.status_code, 200)


class DownloadMIBudgetViewTests(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)
        self.cost_centre_code = 109076
        cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -234567
        self.amount_may = 345216
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.programme_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save

        may_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save

        financial_code_obj1 = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
        )

        may_figure1 = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj1,
            financial_year=year_obj,
        )
        may_figure1.save

        self.year_total = self.amount_apr + self.amount_may

    def test_download_mi_budget(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")
        downloaded_files_url = reverse("download_mi_budget",)

        response = self.factory_get(downloaded_files_url, export_mi_budget_report,)
        self.assertEqual(response.status_code, 302)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()
        response = self.factory_get(downloaded_files_url, export_mi_budget_report,)
        self.assertEqual(response.status_code, 200)
        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True,)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Entity")
        self.assertEqual(ws["A2"].value, "3000")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["H2"].value, self.amount_apr / 100)
        self.assertEqual(ws["I2"].value, self.amount_may / 100)
        self.assertEqual(ws["W2"].value, self.year_total / 100)
        wb.close()

    def test_download_oscar_view(self):
        assert not self.test_user.has_perm("forecast.can_download_oscar")

        downloaded_files_url = reverse("download_oscar_report",)

        # Should have been redirected (no permission)
        with self.assertRaises(PermissionDenied):
            self.factory_get(
                downloaded_files_url, DownloadOscarReturnView,
            )

        can_download_files = Permission.objects.get(codename="can_download_oscar",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        resp = self.factory_get(downloaded_files_url, DownloadOscarReturnView,)

        # Should have been permission now
        self.assertEqual(resp.status_code, 200)
